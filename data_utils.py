###################################################################################
#
# THESIS TITLE: EXTRACTING AND ORGANIZING DISASTER-RELATED PHILIPPINE COMMUNITY
#               RESPONSES FOR AIDING NATIONWIDE RISK REDUCTION PLANNING AND
#               RESPONSE (2020)
#
# AUTHOR/DEVELOPER: Nicco Nocon
# INSTITUTION: De La Salle University, Manila
# EMAIL: noconoccin@gmail.com or nicco_louis_nocon@dlsu.edu.ph
# CONTACT NUMBER: (+63) 917 819 9311
#
# SOURCE OF FUNDING: Philippine-California Advanced Research Institutes (PCARI)
#                    through Commission on Higher Education and Department of
#                    Science and Technology – Science Education Institute
#
###################################################################################
# DATA UTILITIES
# Used for processing Malasakit Responses excel file and text file resources
# > Malasakit: https://opinion.berkeley.edu/pcari/en/landing/
# > Paper: https://ieeexplore.ieee.org/document/8239265
###################################################################################

import malasakit_response
from openpyxl import load_workbook
from nltk.corpus import stopwords
import string


def refresh_excel(filename, protected_cells):
    """
    Function for clearing out values in excel excluding values under a selected protected_cell variable.

    Args:
        filename (str): The file location of the spreadsheet.
        protected_cells (int): The protected cells or number of columns (e.g., 2 columns --> response & tag).
    """
    cell_ctr = 1

    workbook = load_workbook(filename)  # Open a workbook from file
    sheet_names_list = workbook.get_sheet_names()  # Get List of Sheets (to access regardless of sheet name)

    for sheet_name in sheet_names_list:
        sheet = workbook.get_sheet_by_name(sheet_name)  # Access Sheet 1 to remove cells with system output
        if cell_ctr == 1:
            sheet.delete_cols(protected_cells + 1,
                              sheet.max_column)  # Delete columns protected cells onwards for Sheet 1
            cell_ctr += 1
        else:
            sheet.delete_cols(1, sheet.max_column)  # Delete contents for rest of the sheets

    workbook.save(filename)  # Save changes


def read_candidate_excel(filename):
    """
    Function that reads the values in the candidate's excel file and stores the value in a list.

    Args:
        filename (str): The file location of the spreadsheet.

    Returns:
        phrase_list: a list of strings containing the system extracted (phrases) information.
        word_set_list: a list of strings containing the system extracted (word sets) information.
        total_words: a value indicating the total number of words in the input sentence / whole corpus.
    """
    phrase_list = []
    word_set_list = []
    column_list = []

    workbook = load_workbook(filename)  # Open a workbook from file
    sheet_names_list = workbook.get_sheet_names()  # Get List of Sheets (to access regardless of sheet name)

    # Read Phrase Insights (multiple phrases per row)
    sheet = workbook.get_sheet_by_name(sheet_names_list[0])  # Access Sheet 1

    # Word Count
    total_words = 0
    for row in sheet[1:sheet.max_row]:  # Iterates the sheet per row
        cell_value = row[3].value
        # print(cell_value)
        if cell_value is not None:
            if type(cell_value) is str:
                total_words += len(cell_value.split())

    column_index = 5
    # print("Reading Candidate File's:", sheet.title)
    for row in sheet[1:sheet.max_row]:  # Iterates the sheet per row
        for cell in row[column_index:]:  # Iterates the sheet per column
            cell_value = cell.value
            if cell_value is not None:
                if type(cell_value) is str:
                    cell_value = cell_value.strip()
                column_list.append(cell_value)
        phrase_list.append(column_list)
        column_list = []

    # Read Response_ID, Word Insights (one action-target set per row)
    sheet = workbook.get_sheet_by_name(sheet_names_list[1])  # Access Sheet 2
    # print("Reading Candidate File's:", sheet.title)
    for row in sheet[1:sheet.max_row]:  # Iterates the sheet per row
        for cell in row:  # Iterates the sheet per column
            cell_value = cell.value
            if cell_value is not None:
                if type(cell_value) is str:
                    cell_value = cell_value.strip()
                column_list.append(cell_value)
        word_set_list.append(column_list)
        column_list = []

    return phrase_list, word_set_list, total_words


def read_gold_standard_excel(filename):
    """
    Function that reads the values in an excel file and stores the value in a list.

    Args:
        filename (str): The file location of the spreadsheet.

    Returns:
        goldstandard_phrases_list: a list of strings containing the gold standard (phrases) information.
        goldstandard_word_sets_list: a list of strings containing the gold standard (word sets) information.
    """
    goldstandard_phrases_list = []
    goldstandard_word_sets_list = []
    column_list = []

    workbook = load_workbook(filename)  # Open a workbook from file
    sheet_names_list = workbook.get_sheet_names()  # Get List of Sheets (to access regardless of sheet name)

    sheet = workbook.get_sheet_by_name(sheet_names_list[1])  # Access Sheet 1 to get Gold Standard IE Phrases Data
    for row in sheet[1:sheet.max_row]:  # Iterates the sheet per row
        for cell in row:  # Iterates the sheet per column
            cell_value = cell.value
            if cell_value is not None:
                if type(cell_value) is str:
                    cell_value = cell_value.strip()
                column_list.append(cell_value)
        goldstandard_phrases_list.append(column_list)
        column_list = []

    sheet = workbook.get_sheet_by_name(sheet_names_list[2])  # Access Sheet 1 to get Gold Standard IE Phrases Data
    for row in sheet[1:sheet.max_row]:  # Iterates the sheet per row
        for cell in row:  # Iterates the sheet per column
            cell_value = cell.value
            if cell_value is not None:
                if type(cell_value) is str:
                    cell_value = cell_value.strip()
                column_list.append(cell_value)
        goldstandard_word_sets_list.append(column_list)
        column_list = []

    return goldstandard_phrases_list, goldstandard_word_sets_list


def read_clusters_excel(filename, organization_format):
    """
    Function that reads the values in an excel file and stores the value in a list.

    Args:
        filename (str): The file location of the spreadsheet.
        organization_format (str): The organizational format used (i.e., OAE or ORC).

    Returns:
        clusters_list: a list of strings containing the clusters.
        category_count: the number of categories. 0 for OAE.
    """
    clusters_list = []
    response_id = ''
    frequency = ''
    action = ''
    target = ''
    response_ctr = 0
    cell_ctr = 1
    category_count = 0

    workbook = load_workbook(filename)  # Open a workbook from file
    sheet_names_list = workbook.get_sheet_names()  # Get List of Sheets (to access regardless of sheet name)
    sheet = workbook.get_sheet_by_name(sheet_names_list[3])  # Access Sheet 4 to get Ranked Cluster Data

    if organization_format == 'OAE':
        # Iterates the sheet per row
        for row in sheet[1:sheet.max_row]:
            if 'Cluster' not in row[0].value:
                for cell in row:
                    if cell_ctr == 1:
                        response_id = cell.value  # Record response id
                    elif cell_ctr == 2:
                        frequency = cell.value  # Record frequency count
                    elif cell_ctr == 3:
                        action = cell.value  # Record proposed action
                    elif cell_ctr == 4:
                        target = cell.value  # Record target
                    cell_ctr += 1  # Toggle
                # Create Cluster tuple and add into the list
                clusters_list.append((response_id, frequency, action, target))
                cell_ctr = 1  # Reset counter
            else:
                clusters_list.append(row[0].value)

    if organization_format == 'ORC':
        # Iterates the sheet per row
        for row in sheet[1:sheet.max_row]:
            if 'Cluster' not in row[0].value and not row[0].value.isupper():
                for cell in row:
                    if cell_ctr == 1:
                        response_id = cell.value  # Record response id
                    elif cell_ctr == 2:
                        frequency = cell.value  # Record frequency count
                    elif cell_ctr == 3:
                        action = cell.value  # Record proposed action
                    elif cell_ctr == 4:
                        target = cell.value  # Record target
                    cell_ctr += 1  # Toggle
                # Create Cluster tuple and add into the list
                clusters_list.append((response_id, frequency, action, target))
                cell_ctr = 1  # Reset counter
            else:
                clusters_list.append(row[0].value)
                if row[0].value.isupper():
                    category_count += 1

    return clusters_list, category_count


def read_excel(filename):
    """
    Function that reads the values in an excel file and stores the first two columns into the MalasakitResponse object.

    Args:
        filename (str): The file location of the spreadsheet.

    Returns:
        malasakit_response_list: a list of strings containing the Malasakit responses and their respective tags.
    """
    malasakit_response_list = []
    response = ''
    tag = ''
    response_ctr = 0
    cell_ctr = 1

    workbook = load_workbook(filename)  # Open a workbook from file
    sheet_names_list = workbook.get_sheet_names()  # Get List of Sheets (to access regardless of sheet name)
    sheet = workbook.get_sheet_by_name(sheet_names_list[0])  # Access Sheet 1 to get Malasakit Data

    # Iterates the sheet per row
    for row in sheet[1:sheet.max_row]:
        response_ctr += 1  # Response ID
        for cell in row:
            if cell_ctr == 1:
                response = cell.value  # Record response data
            elif cell_ctr == 2:
                tag = cell.value  # Record tag data
            cell_ctr += 1  # Toggle
        # Text cleaning - removed punctuations
        if type(response) is str:
            cleaned_response = response.translate(
                str.maketrans({key: None for key in string.punctuation.replace("-", "")}))
        elif type(response) is int:  # If a response only contains an int number like 0, turn into a string
            cleaned_response = str(response)  # Can also convert into a default value like 'no comment'
        # Create MalasakitResponse object and add into the list
        malasakit_response_list.append(malasakit_response.MalasakitResponse(response_ctr, cleaned_response, tag))
        cell_ctr = 1  # Reset counter

    return malasakit_response_list


def write_excel(filename, malasakit_response_list, clusters_list, ranked_clusters_list):
    """
    Function that writes system output in an excel file.

    Args:
        filename (str): The file location of the spreadsheet.
        malasakit_response_list (list): The list containing MalasakitResponse objects.
        clusters_list (list): The list containing clustered information that was extracted from Malasakit.
        ranked_clusters_list (list): The ranked version of the list containing clustered information.
    """
    workbook = load_workbook(filename)  # Open a workbook from file
    sheet_names_list = workbook.get_sheet_names()  # Get List of Sheets (to access regardless of sheet name)

    # Write POS (stanford format), Response_ID, Phrase Insights (multiple phrases per row)
    sheet = workbook.get_sheet_by_name(sheet_names_list[0])  # Access Sheet 1
    column_index = 3
    # print("Writing on", sheet.title)
    for response in malasakit_response_list:
        sheet.cell(response.response_id, column_index, response.language[0])  # Language
        column_index += 1
        sheet.cell(response.response_id, column_index, response.response)  # Response (normalized)
        column_index += 1
        sheet.cell(response.response_id, column_index, response.fspost_stanford_format)  # POS

        for insight in response.insights_phrase:  # Write Response_ID and Phrase Insights
            column_index += 1
            sheet.cell(response.response_id, column_index, insight)

        column_index = 3  # Reset variable

    # Write Response_ID, Word Insights (one action-target set per row)
    sheet = workbook.get_sheet_by_name(sheet_names_list[1])  # Access Sheet 2
    row_index = 0
    column_index = 0
    # print("Writing on", sheet.title)
    for response in malasakit_response_list:
        for insight in response.insights_words:
            row_index += 1
            for element in insight:
                column_index += 1
                sheet.cell(row_index, column_index, element)
            column_index = 0  # Reset Column

    # Write Cluster/Entry number, Appended Response_IDs, Frequency Count, Actions, Targets
    sheet = workbook.get_sheet_by_name(sheet_names_list[2])  # Access Sheet 3
    row_index = 1
    column_index = 1
    cluster_ctr = 0
    # print("Writing on", sheet.title)
    for cluster in clusters_list:  # Traverse by row
        if isinstance(cluster, str):  # 'If element is a str == true' means that index 0 is the category string not id
            sheet.cell(row_index, column_index, cluster.upper())
            cluster_ctr = 0  # Reset index
            row_index += 1
        else:
            cluster_ctr += 1  # Cluster 1, ...
            sheet.cell(row_index, column_index, "Cluster " + str(cluster_ctr))
            row_index += 1

            for element in cluster:  # Print in column for each element
                sheet.cell(row_index, column_index, element)
                column_index += 1
            column_index = 1  # Reset Column
            row_index += 1

    # Write Cluster/Entry number, Response_IDs, Ranked Frequencies, Actions, Targets
    sheet = workbook.get_sheet_by_name(sheet_names_list[3])  # Access Sheet 4
    row_index = 1
    column_index = 1
    cluster_ctr = 0
    # print("Writing on", sheet.title)
    for cluster in ranked_clusters_list:  # Traverse by row
        if isinstance(cluster, str):  # 'If element is a str == true' means that index 0 is the category string not id
            sheet.cell(row_index, column_index, cluster.upper())
            cluster_ctr = 0  # Reset index
            row_index += 1
        else:
            cluster_ctr += 1  # Cluster 1, ...
            sheet.cell(row_index, column_index, "Cluster " + str(cluster_ctr))
            row_index += 1

            for element in cluster:  # Print in column for each element
                sheet.cell(row_index, column_index, element)
                column_index += 1
            column_index = 1  # Reset Column
            row_index += 1

    workbook.save(filename)  # Save changes


def text_to_list(filename):
    """
    Function for transforming a given text file into a list of list [s1[w1,w2,...,wN], ..., sN[w1,w2,...,wN]].

    Args:
        filename (str): The file location of the text file.

    Returns:
        sentence_list: a list of strings containing sentences found on the file.
    """
    sentence_list = []
    file = open(filename)
    for line in file:
        sentence_list.append(line.split())  # Returns a white-spaced split list of words and store in sentence_list
    file.close()
    return sentence_list


def text_to_list_without_stopwords(filename):
    """
    Function for transforming a given text file into a list of list removing tl/en stopwords in the process.

    Args:
        filename (str): The file location of the text file.

    Returns:
        sentence_list: a list of strings containing sentences found on the file (without stopwords).
    """
    sentence_list = []
    tl_stop_words_list = get_stopwords_from_file('model/dictionary/tl_stopwords.txt')
    en_stop_words_list = list(stopwords.words('english'))  # Uses nltk stopwords
    stopwords_list = tl_stop_words_list + en_stop_words_list
    file = open(filename)
    for line in file:
        words_list = line.split()
        filtered_sentence = [word for word in words_list if word not in stopwords_list]
        sentence_list.append(filtered_sentence)  # Returns a white-spaced split list of words and store in sentence_list
    file.close()
    return sentence_list


def write_text_file(filename, string_list):
    """
    Function that writes the strings in a list to a text file. Normally to be read by the normalizer module.

    Args:
        filename (str): The file location of the text file.
        string_list: The list to be written in the text file.
    """
    file = open(filename, 'w+')
    for element in string_list:
        file.write(element + '\n')  # Write per element/line in the list
    file.close()


def read_text_file(filename):
    """
    Function that reads the strings in a file and transfer them into a list.

    Args:
        filename (str): The file location of the text file.

    Returns:
        string_list: a list containing the contents of the text file.
    """
    file = open(filename, 'r')
    string_list = file.readlines()  # Put in the list the string contents per new line
    string_list = list(map(lambda sentence: sentence.strip(), string_list))  # Remove \n
    file.close()
    return string_list


def get_stopwords_from_file(filename):
    """
    Function for transforming stop words in a given file to a list. Can use read_text_file instead.

    Args:
        filename (str): The file location of the text file.

    Returns:
        stopwords_list: a list containing the stopwords found on the text file.
    """
    stopwords_list = []
    file = open(filename)
    for line in file:
        stopwords_list.append(line.rstrip())
    file.close()
    return stopwords_list
